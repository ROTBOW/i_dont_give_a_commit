import os
import re
from collections import defaultdict as ddict
from datetime import date, datetime
from time import sleep, time

import requests
from alive_progress import alive_bar
from bs4 import BeautifulSoup as bs
from openpyxl import load_workbook

DIR = os.path.dirname(os.path.realpath(__file__))
TARGET = fr'{DIR}\\target'
RES = fr'{DIR}\\res'

class IDGAC:
    
    def __init__(self) -> None:
        """
        This is a constructor function that initializes three data structures - a nested dictionary,
        another nested dictionary, and a list.
        """
        self.seekers_by_coach = ddict(dict)
        self.lacking_by_coach = ddict(dict)
        self.seekers_no_link = list()
        
        
    def __grab_data_from_file(self) -> None:
        """
        This function reads data from an Excel file in the target folder and populates a dictionary with the data.
        """
        target_file = os.listdir(TARGET)[0]
        data = load_workbook(fr'{TARGET}\\{target_file}')
        sheet = data.active
        cells = {
            0: 'seeker',
            1: 'coach',
            2: 'url'
        }

        with alive_bar(sheet.max_row-1, title="Grabing Data...") as bar:
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                curr_row = dict()
                for idx, cell in enumerate(row):
                    curr_row[cells[idx]] = cell.value
                        
                
                coach = curr_row['coach'] if curr_row['coach'] != ' ' else 'Placements'
                
                self.seekers_by_coach[coach][curr_row['seeker']] = curr_row['url']
                bar()
    
    
    def __prune_linkless(self) -> None:
        """
        This function prunes linkless seekers from a dictionary and writes their names to a text file.
        """
        if not self.seekers_by_coach:
            raise Exception('NO DATA - Are you sure the file was parsed properly?')
        
        linkless = set()
        
        with alive_bar(0, title="Pruning linkless seekers...") as bar:
            for coach in self.seekers_by_coach:
                ll = set()
                for seeker, url in self.seekers_by_coach[coach].items():
                    if url == '':
                        ll.add(seeker)
                    bar()
        
                for seeker in ll:
                    self.seekers_by_coach[coach].pop(seeker)
                    bar()
                    
                linkless.update(ll)
        
        with open(f'{RES}\linkless_seekers_{date.today()}.txt', 'w') as f:
            f.write('\n'.join(sorted(linkless)))
        
    def __last_seven_days(self, commit_date: str, commit_month: str, commit_year: str) -> bool:
        """
        This function checks if a commit was made within the last seven days.
        
        :param commit_date: The day of the month on which a commit was made (as a string)
        :type commit_date: str
        :param commit_month: The month in which a commit was made, represented as a string (e.g.
        "January", "February", etc.)
        :type commit_month: str
        :param commit_year: The year in which a commit was made
        :type commit_year: str
        :return: a boolean value indicating whether the commit date provided as input is within the last
        seven days from the current date.
        """
        today = date.today().day
        this_month = date.today().month
        this_year = date.today().year
        
        commit_month = datetime.strptime(commit_month, '%B').month
        commit_date = int(commit_date)
        commit_year = int(commit_year)
        
        return all([
            today-7 <= commit_date <= today,
            commit_month == this_month,
            commit_year == this_year
            ])
             
        
        
    def __get_commits(self, url: str) -> int:
        """
        This function retrieves the number of commits made by a user in the last seven days from a given
        GitHub URL.
        
        :param url: The URL of a user's GitHub page
        :type url: str
        :return: an integer value representing the number of commits made in the last seven days on a
        given GitHub repository.
        """
        sleep(.5)
        res = requests.get(url)
        soup = bs(res.text, 'html.parser')
        days = list(day.text for day in soup.find_all('rect', {'class': 'ContributionCalendar-day'}) if day.text)
        commits = 0
        
        for day in days:
            match = re.match(r"^(?P<count>(No|\d+)) contribution[s]? on \w+, (?P<month>\w+) (?P<date>\d{1,2}), (?P<year>\d{4})", day)
            if match:
                if self.__last_seven_days(match.group('date'), match.group('month'), match.group('year')):
                    commits += int(match.group('count')) if match.group('count') != 'No' else 0
            else:
                raise Exception(f'REGEX FAILED - Check if I\'m just a big dumb, but it should be able to find a match for all days\nIf anything here is the string it failed to match: {day}')
        
        
        return commits
    
    def __linear_get_seeker_commits(self) -> None:
        """
        This function iterates through each seeker of each coach and gets their commits, storing the
        ones with less than 5 commits in a dictionary.
        """
        
        # you can change below to loop over all the coaches you want to get the seeker commits for
        # if you use this line:
            # for coach in self.seekers_by_coach:
        # it will go over all coachs    
        
        for coach in ['Josiah Leon', 'Peter Joh']:
            with alive_bar(len(self.seekers_by_coach[coach]), title=f"Getting {coach}'s seekers' commits") as bar:
                for seeker in self.seekers_by_coach[coach]:
                    commits = self.__get_commits(self.seekers_by_coach[coach][seeker])
                    if commits <= 4:
                        self.lacking_by_coach[coach][seeker] = commits
                    bar()
        
    def __res_to_file(self) -> None:
        """
        This function writes a report of lacking commits by coach and seeker to a file.
        """
        with open(f'{RES}\lacking_commits_{date.today()}.txt', 'w') as f:
            for coach in self.lacking_by_coach:
                f.write(coach + '\n')
                for seeker, commits in self.lacking_by_coach[coach].items():
                    f.write(f'    {seeker}: {commits}\n')
                f.write('\n')
    
    def main(self):
        """
        The main function grabs data from a file, prunes linkless data, gets commits using linear
        seeking, and writes the results to a file.
        """
        self.__grab_data_from_file()
        self.__prune_linkless()
        # Func below can be swapped with a theading ver for more speed if needed
        self.__linear_get_seeker_commits()
        self.__res_to_file()
    
    
if __name__ == '__main__':
    if not os.path.isdir(RES):
        os.mkdir(RES)
    start = time()
    idgac = IDGAC()
    idgac.main()
    end = time()
    
    print(f'Total Time: {round(end-start, 2)}s')